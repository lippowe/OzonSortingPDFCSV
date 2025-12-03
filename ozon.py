import streamlit as st
import pandas as pd
import re
import io
from pypdf import PdfReader, PdfWriter
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

FBS_PREFIXES = {
    "Озон": "204514",
    "Рига": "2503733",
    "Плутон": "3021812"
}


def extract_order_number_prefix(order_string):
    """Извлекает префикс номера заказа."""
    if not isinstance(order_string, str):
        order_string = str(order_string)
    match = re.search(r'^(\d+)-', order_string)
    if match:
        return match.group(1)
    else:
        return None


# Функция extract_sticker_from_order была удалена, так как не использовалась.


def sort_dataframe(df):
    """Сортирует DataFrame в соответствии с заданными приоритетами."""
    required_cols = ['Артикул', 'Количество', 'Наименование товара', 'Номер отправления', 'Стикер']
    for col in required_cols:
        if col not in df.columns:
            df[col] = ''

    df['Количество'] = pd.to_numeric(df['Количество'], errors='coerce').fillna(0)
    original_article_case = df['Артикул'].astype(str).copy()  # Use .copy() to avoid SettingWithCopyWarning
    df['Артикул_lower'] = df['Артикул'].astype(str).str.lower()
    df['Наименование товара_lower'] = df['Наименование товара'].astype(str).str.lower()

    def get_article_core(article):
        """Извлекает основную часть артикула, убирая суффиксы."""
        # This regex looks for a single letter followed by digits at the end
        match = re.search(r'([a-z]\d+)$', article)
        if match:
            end_of_core = match.start()
            return article[:end_of_core].strip()
        else:
            return article.strip()

    df['article_core'] = df['Артикул_lower'].apply(get_article_core)
    core_counts = df['article_core'].value_counts()
    df['core_repeat_count'] = df['article_core'].map(core_counts)
    sticker_counts = df['Артикул_lower'].value_counts()
    df['full_sticker_repeat_count'] = df['Артикул_lower'].map(sticker_counts)

    df['shipment_sticker_key'] = df['Номер отправления'].astype(str) + '_' + df['Стикер'].astype(str)
    shipment_sticker_counts = df['shipment_sticker_key'].value_counts()
    df['shipment_sticker_repeated'] = df['shipment_sticker_key'].map(shipment_sticker_counts)
    df['shipment_sticker_repeated_flag'] = df['shipment_sticker_repeated'] > 1

    df['has_k_prefix_num'] = df['Артикул_lower'].str.contains(r'.*[k][2-5]\d*.*', na=False)
    df['qty_greater_than_1'] = df['Количество'] > 1
    df['article_repeated'] = df['full_sticker_repeat_count'] > 1

    df['name_article_key'] = df['Наименование товара_lower'].astype(str) + '_' + df['Артикул_lower'].astype(str)
    name_article_counts = df['name_article_key'].value_counts()
    df['name_article_repeated'] = df['name_article_key'].map(name_article_counts)

    df['k_num_suffix'] = 0
    # Единое регулярное выражение: k, за которым следуют 2-5, а затем любые цифры в конце
    k_match = df['Артикул_lower'].str.extract(r'.*[k]([2-5]\d*)$', expand=False)
    df['k_num_suffix'] = pd.to_numeric(k_match, errors='coerce').fillna(0)

    df['sort_level'] = 4.0
    priority1_mask = (df['core_repeat_count'] > 1) & (df['has_k_prefix_num'])
    df.loc[priority1_mask, 'sort_level'] = 1.0
    priority2_mask = (df['full_sticker_repeat_count'] > 1) & (df['qty_greater_than_1']) & (df['sort_level'] == 4.0)
    df.loc[priority2_mask, 'sort_level'] = 2.0
    priority3_mask = (df['full_sticker_repeat_count'] > 1) & (df['sort_level'] == 4.0)
    df.loc[priority3_mask, 'sort_level'] = 3.0

    df = df.sort_values(
        by=[
            'shipment_sticker_repeated_flag',  # Приоритет 1: Повторение "Номер отправления" и "Стикер" (True first)
            'has_k_prefix_num',  # Приоритет 2: Наличие k/K с числом (True first)
            'k_num_suffix',  # Сортировка по номеру после k/K (убывание)
            'qty_greater_than_1',  # Приоритет 3: Количество > 1 (True first)
            'article_repeated',  # Приоритет 4: Повторяющийся артикул (True first)
            'name_article_repeated',  # Приоритет 5: Повторение "Наименование товара" и "Артикул" (убывание)
            'sort_level',  # Остальные критерии (по возрастанию, P1 -> P2 -> P3 -> P4)
            'article_core',  # Дополнительные критерии для стабильной сортировки
            'core_repeat_count',  # Дополнительные критерии
            'Количество',  # Дополнительные критерии
            'Наименование товара_lower',  # Дополнительные критерии
            'Артикул_lower'  # Дополнительные критерии
        ],
        ascending=[
            False,  # 'shipment_sticker_repeated_flag': Сначала True (повторяется)
            False,  # 'has_k_prefix_num': Сначала True (есть k/K)
            False,  # 'k_num_suffix':  Убывание (сначала больше)
            False,  # 'qty_greater_than_1': Сначала True (Количество > 1)
            False,  # 'article_repeated': Сначала True (повторяется)
            False,  # 'name_article_repeated': По убыванию (сначала больше)
            True,  # 'sort_level':  По возрастанию
            True,  # 'article_core'
            False,  # 'core_repeat_count' (больше повторов - выше приоритет)
            False,  # 'Количество' (больше - выше)
            True,  # 'Наименование товара_lower'
            True  # 'Артикул_lower'
        ]
    )

    df['Артикул'] = original_article_case
    return df


def extract_sticker_data_from_pdf(pdf_file, fbs_prefix):
    """Извлекает данные стикеров из PDF."""
    sticker_data = {}
    try:
        reader = PdfReader(pdf_file)
        for page_num, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                pattern = r"FBS:\s*" + re.escape(fbs_prefix) + r"[\s\S]*?(\d+)-"
                match = re.search(pattern, text)

                if match:
                    sticker_number = match.group(1)
                    sticker_data[page_num + 1] = sticker_number  # Store as {page_number: sticker_value}
                # Removed: st.write(r"Номер " + sticker_number) to reduce UI clutter
    except Exception as e:
        st.error(f"Ошибка при обработке PDF файла: {e}")
    return sticker_data


def reorder_pdf_pages(pdf_file, page_order_mapping):
    """Переупорядочивает страницы PDF."""
    try:
        reader = PdfReader(pdf_file)
        writer = PdfWriter()

        original_pages = {i + 1: page for i, page in enumerate(reader.pages)}

        # Validate all pages requested in mapping exist in original PDF
        for original_page_num, sticker_value in page_order_mapping:
            if original_page_num not in original_pages:
                st.warning(
                    f"Страница {original_page_num} (стикер: {sticker_value}) не найдена в исходном PDF. Она будет пропущена.")
                continue

        for original_page_num, _ in page_order_mapping:
            if original_page_num in original_pages:
                page_to_add = original_pages[original_page_num]
                writer.add_page(page_to_add)
        return writer
    except Exception as e:
        st.error(f"Ошибка при переупорядочивании страниц PDF: {e}")
        return None


def get_last_4_digits(value):
    """Извлекает последние 4 цифры из значения."""
    if pd.isna(value):
        return ""
    value_str = str(value)
    match = re.search(r'(\d{4})$', value_str)
    if match:
        return match.group(0)
    else:
        digits_only = "".join(filter(str.isdigit, value_str))
        if len(digits_only) >= 4:
            return digits_only[-4:]
        else:
            return ""


def _apply_excel_formatting(sheet, df_data, sheet_name, fbs_option, num_pdf_pages=None):
    """Применяет общую стилизацию к листу Excel."""
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='left', vertical='top')

    # Add header info
    sheet['B1'] = f'{sheet_name} OZON'
    sheet['B1'].font = Font(bold=True, size=16)

    sheet['B2'] = f'Склад: {fbs_option}'
    sheet['B2'].font = Font(bold=True, size=13)

    sheet['B3'] = 'Дата: ' + datetime.now().strftime("%Y-%m-%d %H:%M")
    sheet['B3'].font = Font(bold=True, size=13)

    if num_pdf_pages is not None:
        sheet['B4'] = f'Количество отправлений: {num_pdf_pages}'
        sheet['B4'].font = Font(bold=True, size=13)

    # Apply header formatting
    for col_num in range(1, df_data.shape[1] + 1):
        cell = sheet.cell(row=6, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment

    # Apply data formatting and conditional bolding for quantity
    for row_num in range(7, sheet.max_row + 1):
        for col_num in range(1, df_data.shape[1] + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = data_alignment

            if sheet.cell(row=6, column=col_num).value == 'Кол-во':
                if isinstance(cell.value, (int, float)) and cell.value > 1:
                    cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col_num in range(1, df_data.shape[1] + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        # Check header and data rows for max length
        for row_num in range(6, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        sheet.column_dimensions[column_letter].width = max_length + 2  # Add padding

    # Printing settings
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.paperSize = 9  # A4
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0
    sheet.page_setup.fitToWidth = 1  # Fit to one page width
    sheet.page_setup.fitToHeight = 0  # No height limit
    sheet.freeze_panes = 'A7'  # Freeze header row


def create_main_excel_file(df_main, fbs_option, num_pdf_pages):
    """Создает Excel файл для основного листа подбора."""
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        sheet_name_main = 'Лист подбора'
        df_main.to_excel(writer, sheet_name=sheet_name_main, index=False, startrow=5)
        sheet_main = writer.sheets[sheet_name_main]
        _apply_excel_formatting(sheet_main, df_main, sheet_name_main, fbs_option, num_pdf_pages)
    excel_buffer.seek(0)
    return excel_buffer


def create_repeats_excel_file(df_repeats, fbs_option):
    """Создает Excel файл для листа с повторяющимися заказами."""
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        repeats_sheet_name = 'Повторы'
        df_repeats.to_excel(writer, sheet_name=repeats_sheet_name, index=False, startrow=5)
        sheet_repeats = writer.sheets[repeats_sheet_name]
        # For repeats, num_pdf_pages is not relevant for the header
        _apply_excel_formatting(sheet_repeats, df_repeats, repeats_sheet_name, fbs_option, num_pdf_pages=None)
    excel_buffer.seek(0)
    return excel_buffer


def read_csv_with_encoding(uploaded_csv_file):
    """
    Пытается прочитать CSV файл с разными кодировками и определяет столбец 'Наименование товара'.
    """
    encodings_to_try = ['utf-8', 'cp1251', 'latin1']
    sep_options = [';', ',', '\t']
    possible_name_columns = ['Наименование товара', 'Название товара', 'Название']

    uploaded_csv_file.seek(0)  # Reset file pointer for reliable reading

    for sep in sep_options:
        for encoding in encodings_to_try:
            try:
                uploaded_csv_file.seek(0)
                df = pd.read_csv(uploaded_csv_file, sep=sep, encoding=encoding)

                name_column = None
                for col in possible_name_columns:
                    if col in df.columns:
                        name_column = col
                        break

                if name_column is None:
                    continue  # Try next encoding/separator

                if name_column != 'Наименование товара':
                    df = df.rename(columns={name_column: 'Наименование товара'})
                return df  # Successfully read and renamed, return df

            except (UnicodeDecodeError, pd.errors.ParserError):
                pass  # Try next encoding/separator
            except Exception as e:
                st.write(f"Другая ошибка при чтении с разделителем '{sep}' и кодировкой '{encoding}': {e}")
                pass  # Catch other potential errors

    st.error(
        f"Не удалось прочитать CSV файл ни с одной из предложенных кодировок/разделителей ({', '.join(encodings_to_try)} / {', '.join(sep_options)}). "
        f"Проверены столбцы: {', '.join(possible_name_columns)}. Пожалуйста, убедитесь, что файл корректный и содержит нужные столбцы.")

    # Final attempt with default pandas settings
    try:
        uploaded_csv_file.seek(0)
        df = pd.read_csv(uploaded_csv_file)
        name_column = None
        for col in possible_name_columns:
            if col in df.columns:
                name_column = col
                break
        if name_column is not None:
            if name_column != 'Наименование товара':
                df = df.rename(columns={name_column: 'Наименование товара'})
            st.info("Файл прочитан с использованием стандартных настроек pandas.")
            return df
    except Exception as e:
        st.error(f"Финальная попытка чтения с помощью стандартных настроек pandas не удалась: {e}")

    return None


def main():
    """Основная логика приложения Streamlit."""
    st.set_page_config(layout="wide")
    st.title("Обработка заказов Озон: PDF и CSV")

    fbs_option = st.selectbox("Выберите тип FBS", list(FBS_PREFIXES.keys()))
    fbs_prefix = FBS_PREFIXES[fbs_option]

    st.header("1. Загрузка файлов")
    uploaded_csv_file = st.file_uploader("Загрузите CSV файл с заказами", type=["csv", "txt"])
    uploaded_pdf_file = st.file_uploader("Загрузите PDF файл со стикерами", type="pdf")

    if uploaded_csv_file and uploaded_pdf_file:
        st.success("Файлы успешно загружены! Приступаю к обработке...")

        try:
            df_original = read_csv_with_encoding(uploaded_csv_file)

            if df_original is None:
                st.stop()

            df_original['Наименование товара'] = df_original['Наименование товара'].astype(str).fillna('')

            if 'Номер заказа' not in df_original.columns:
                st.error("В CSV файле отсутствует обязательный столбец 'Номер заказа'.")
                st.stop()

            df_original['Стикер'] = df_original['Номер заказа'].apply(extract_order_number_prefix)
            df_with_order_prefix = df_original.dropna(subset=['Стикер']).copy()

            if df_with_order_prefix.empty:
                st.warning(
                    "Не найдено ни одного номера заказа в формате 'число-' в колонке 'Номер заказа' CSV файла. Проверьте формат номеров заказов.")
                st.stop()
            else:
                df_sorted = sort_dataframe(df_with_order_prefix)
                df_sorted = df_sorted.reset_index(drop=True)

                df_repeats = df_sorted[df_sorted['shipment_sticker_repeated_flag']].copy()
                df_main_sheet = df_sorted[~df_sorted['shipment_sticker_repeated_flag']].copy()

                num_rows_main = len(df_main_sheet)
                df_main_sheet['Код'] = pd.Series(range(1, num_rows_main + 1), index=df_main_sheet.index)

                start_num_repeats = df_main_sheet['Код'].max() + 1 if not df_main_sheet.empty else 1
                num_rows_repeats = len(df_repeats)
                df_repeats['Код'] = pd.Series(range(start_num_repeats, start_num_repeats + num_rows_repeats),
                                              index=df_repeats.index)

                df_main_sheet = df_main_sheet.rename(columns={'Количество': 'Кол-во'})
                df_repeats = df_repeats.rename(columns={'Количество': 'Кол-во'})

                desired_columns_excel = ['Код', 'Номер отправления', 'Наименование товара', 'Артикул', 'Кол-во',
                                         'Стикер']

                df_for_excel_main = df_main_sheet[desired_columns_excel].copy()
                df_for_excel_repeats = df_repeats[desired_columns_excel].copy()

                # Применяем извлечение последних 4 цифр к колонке 'Стикер' для обоих DataFrame'ов
                df_for_excel_main['Стикер'] = df_for_excel_main['Стикер'].apply(get_last_4_digits)
                df_for_excel_repeats['Стикер'] = df_for_excel_repeats['Стикер'].apply(get_last_4_digits)

                # Логика для скрытия повторяющихся стикеров в листе "Повторы"
                if not df_for_excel_repeats.empty:
                    df_for_excel_repeats['Sticker_Group'] = (
                                df_for_excel_repeats['Стикер'] != df_for_excel_repeats['Стикер'].shift()).cumsum()
                    df_for_excel_repeats.loc[df_for_excel_repeats.duplicated(subset=['Стикер', 'Sticker_Group'],
                                                                             keep='first'), 'Стикер'] = ''
                    df_for_excel_repeats.drop(columns=['Sticker_Group'], inplace=True)

                # Извлекаем данные стикеров из PDF
                pdf_sticker_data = extract_sticker_data_from_pdf(uploaded_pdf_file, fbs_prefix)

                reader = PdfReader(uploaded_pdf_file)
                num_pdf_pages_original = len(reader.pages)  # Общее количество страниц в исходном PDF

                if not pdf_sticker_data:
                    st.warning(
                        f"Не удалось извлечь ни одного стикера из PDF файла. Проверьте, соответствует ли формат стикера шаблону 'FBS: {fbs_prefix} XXXXX-'.")

                    # Предлагаем скачать Excel файлы даже без переупорядоченного PDF
                    st.header("- Лист подбора (Excel) -")
                    main_excel_buffer = create_main_excel_file(df_for_excel_main, fbs_option,
                                                               0)  # 0, т.к. стикеры не найдены
                    st.download_button(
                        label="Скачать основной Excel файл",
                        data=main_excel_buffer.getvalue(),
                        file_name=f"Ozon_Main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    if not df_for_excel_repeats.empty:
                        st.header("- Повторы (Excel) -")
                        repeats_excel_buffer = create_repeats_excel_file(df_for_excel_repeats, fbs_option)
                        st.download_button(
                            label="Скачать Excel файл с повторами",
                            data=repeats_excel_buffer.getvalue(),
                            file_name=f"Ozon_Repeats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    st.stop()
                else:
                    pdf_pages_in_desired_order = []
                    missing_csv_stickers_in_pdf = set()

                    temp_pdf_sticker_data = pdf_sticker_data.copy()

                    # Объединяем DataFrame для определения окончательного порядка PDF
                    # (используем полный префикс стикера для сопоставления с PDF)
                    df_combined_for_pdf_order = pd.concat([
                        df_main_sheet[['Стикер', 'Код']],
                        df_repeats[['Стикер', 'Код']]
                    ], ignore_index=True).sort_values(by='Код')

                    for index, row in df_combined_for_pdf_order.iterrows():
                        csv_identifier = row['Стикер']  # Это полный префикс стикера
                        found_page = None

                        for page_num, pdf_sticker_value in list(temp_pdf_sticker_data.items()):
                            if pdf_sticker_value == csv_identifier:
                                found_page = (page_num, pdf_sticker_value)
                                del temp_pdf_sticker_data[page_num]  # Помечаем как использованный
                                break

                        if found_page:
                            pdf_pages_in_desired_order.append(found_page)
                        else:
                            missing_csv_stickers_in_pdf.add(csv_identifier)

                    if missing_csv_stickers_in_pdf:
                        st.warning(
                            f"Следующие стикеры из CSV (после сортировки) не найдены в PDF: "
                            f"{', '.join(sorted(list(missing_csv_stickers_in_pdf)))}."
                        )

                    if temp_pdf_sticker_data:
                        unused_pdf_stickers_values = sorted(list(set(temp_pdf_sticker_data.values())))
                        st.info(
                            f"В PDF файле найдены стикеры, не соответствующие заказам в CSV: "
                            f"{', '.join(unused_pdf_stickers_values)}."
                        )

                    num_mapped_pdf_pages = len(pdf_pages_in_desired_order)
                    if num_mapped_pdf_pages == 0:
                        st.error("После сопоставления не осталось ни одной страницы для переупорядочивания PDF.")
                        # Предлагаем скачать Excel файлы даже без переупорядоченного PDF
                        st.header("- Лист подбора (Excel) -")
                        main_excel_buffer = create_main_excel_file(df_for_excel_main, fbs_option, 0)
                        st.download_button(
                            label="Скачать основной Excel файл",
                            data=main_excel_buffer.getvalue(),
                            file_name=f"Ozon_Main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        if not df_for_excel_repeats.empty:
                            st.header("- Повторы (Excel) -")
                            repeats_excel_buffer = create_repeats_excel_file(df_for_excel_repeats, fbs_option)
                            st.download_button(
                                label="Скачать Excel файл с повторами",
                                data=repeats_excel_buffer.getvalue(),
                                file_name=f"Ozon_Repeats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        st.stop()

                    reordered_pdf_writer = reorder_pdf_pages(uploaded_pdf_file, pdf_pages_in_desired_order)

                    if reordered_pdf_writer:
                        st.success("Стикеры успешно переупорядочены!")

                        st.header("--- Скачать результаты ---")

                        # --- Скачивание Excel файлов ---
                        col1, col2 = st.columns(2)
                        with col1:
                            st.subheader("Лист подбора (Excel)")
                            main_excel_buffer = create_main_excel_file(df_for_excel_main, fbs_option,
                                                                       num_mapped_pdf_pages)
                            if main_excel_buffer:
                                st.download_button(
                                    label="Скачать основной Excel файл",
                                    data=main_excel_buffer.getvalue(),
                                    file_name=f"Ozon_Main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                        with col2:
                            st.subheader("Соеденнёные заказы (Excel)")
                            if not df_for_excel_repeats.empty:
                                repeats_excel_buffer = create_repeats_excel_file(df_for_excel_repeats, fbs_option)
                                if repeats_excel_buffer:
                                    st.download_button(
                                        label="Скачать Excel файл с повторами",
                                        data=repeats_excel_buffer.getvalue(),
                                        file_name=f"Ozon_Repeats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            else:
                                st.info("Нет повторяющихся заказов для отдельного Excel файла.")

                        # --- Скачивание PDF ---
                        st.subheader("Переупорядоченные стикеры (PDF)")
                        pdf_output_buffer = io.BytesIO()
                        reordered_pdf_writer.write(pdf_output_buffer)
                        pdf_output_buffer.seek(0)

                        st.download_button(
                            label="Скачать PDF стикеры",
                            data=pdf_output_buffer,
                            file_name=f"Ozon_Stickers-{datetime.now().strftime('%H-%M-%S')}.pdf",
                            mime="application/pdf"
                        )
                    else:
                        st.error("Не удалось переупорядочить PDF. Пожалуйста, проверьте логи вверху.")

        except Exception as e:
            st.error(f"Произошла непредвиденная ошибка при обработке файлов: {e}")
            st.exception(e)  # Показать полный traceback для отладки


if __name__ == "__main__":
    main()

